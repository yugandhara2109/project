from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import Combobox
from PIL import Image, ImageTk
import openpyxl
from openpyxl import Workbook
import pathlib
import os
from datetime import date
import tkinter as tk


# Dummy credentials (You can replace this with a database check)
USERNAME = "admin"
PASSWORD = "password"

def login():
    """Function to validate login credentials"""
    user = username_entry.get()
    passw = password_entry.get()
    
    if user == USERNAME and passw == PASSWORD:
        messagebox.showinfo("Login Success", "Welcome to Student Registration System!")
        login_window.destroy()  # Close login window
        open_registration_window()  # Open the main app
    else:
        messagebox.showerror("Login Failed", "Invalid username or password")

def open_registration_window():
    """Function to open the Student Registration System window"""
    import student_registration  # Import the main registration script

# Creating login window
login_window = Tk()
login_window.title("Login")
login_window.geometry("400x300")
login_window.config(bg="#06283D")

Label(login_window, text="Login", font="Arial 20 bold", bg="#06283D", fg="white").pack(pady=20)

Label(login_window, text="Username", font="Arial 14", bg="#06283D", fg="white").pack()
username_entry = Entry(login_window, font="Arial 14", width=25)
username_entry.pack(pady=5)

Label(login_window, text="Password", font="Arial 14", bg="#06283D", fg="white").pack()
password_entry = Entry(login_window, font="Arial 14", width=25, show="*")
password_entry.pack(pady=5)

Button(login_window, text="Login", font="Arial 14", command=login, bg="#68ddfa", width=15).pack(pady=20)

login_window.mainloop()

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

root = tk.Tk()
root.title("Student Registration System")
root.geometry("1000x600+100+50")
root.config(bg=background)

# Check if the Excel file exists, if not, create one
file = pathlib.Path('Student_data.xlsx')
if not file.exists():
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date of Registration"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father Name"
    sheet['J1'] = "Mother Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"
    file.save('Student_data.xlsx')


# Exit
def Exit():
    root.destroy()


# Show Image
def showimage():
    global filename
    global img

    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select image file",
                                          filetypes=[("JPG File", "*.jpg"),
                                                    ("PNG File", "*.png"),
                                                    ("All Files", ".")])

    if filename:  # Proceed if a valid file is selected
        try:
            img = Image.open(filename)
            resized_image = img.resize((190, 190))  # Resize the image to fit the space
            photo2 = ImageTk.PhotoImage(resized_image)

            lbl.config(image=photo2)
            lbl.image = photo2  # Keep a reference to avoid garbage collection
        except Exception as e:
            messagebox.showerror("Error", f"Error loading image: {e}")
    else:
        messagebox.showwarning("Warning", "No file selected!")


# Registration Number
def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set("1")


# Clear
def Clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skills.set('')
    F_Name.set('')
    M_Name.set('')
    F_Occupation.set('')
    M_Occupation.set('')
    Class.set("Select Class")

    registration_no()
    saveButton.config(state='normal')

    # Reset image to a default image
    try:
        img1 = Image.open('Images/sear.png')  # Ensure this image exists in your directory
        img1 = img1.resize((190, 190))  # Resize to fit
        photo2 = ImageTk.PhotoImage(img1)

        lbl.config(image=photo2)
        lbl.image = photo2
    except Exception as e:
        print("Error loading default image:", e)

    img = ""


# Save
def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("Error", "Select Gender")

    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skills.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = F_Occupation.get()
    M1 = M_Occupation.get()

    if N1 == "" or C1 == "Select Class" or D2 == "" or Re1 == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("Error", "Few Data is missing")
    else:
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=Re1)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=fathername)
        sheet.cell(column=10, row=sheet.max_row, value=mothername)
        sheet.cell(column=11, row=sheet.max_row, value=F1)
        sheet.cell(column=12, row=sheet.max_row, value=M1)
        file.save('Student_data.xlsx')

        try:
            if img:
                img.save(f"Images/{R1}.jpg")
        except:
            messagebox.showinfo("Info", "Profile picture is not available")

        messagebox.showinfo("Info", "Successfully data entered")
        Clear()
        registration_no()
#search
def search():
    text=Search.get()
    Clear()
    saveButton.config(state='disable')

    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active
    
    for row in sheet.rows:
        if row[0].value == int(text):
            name=row[0]
#            print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
#            print(reg_no_position)
#           print(reg_number)


    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid registration number")

    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value
    

    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)

    if x4=='Female':
        R1.select()
    else:
        R1.select()

    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skills.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    F_Occupation.set(x11)
    M_Occupation.set(x12)

    img = (Image.open("Images/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2



# Gender Selection
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"


# UI Elements and Layout
Label(root, text="Email: yugandhara@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="#c36464", fg='#fff', font='arial 30 bold').pack(side=TOP, fill=X)

#Search box to update
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font="arial 20").place(x=680, y=180)
imageicon3 = PhotoImage(file="Images/sear.png")
Srch = Button(root, text="Search", compound=LEFT, width=20, bg='#68ddfa', font="arial 13 bold", command=search)
Srch.place(x=950, y=180)

#imageicon4=PhotoImage(file="Images/sear.png")
#Update_button=Button(root,imageicon4,bg="#c36464",command=Update)
#Update_button.place(x=110,y=66)

Label(root, text="Registration No.:", font="arial 17", fg=framebg, bg=background).place(x=200, y=270)
Label(root, text="Date:", font="arial 17", fg=framebg, bg=background).place(x=800, y=270)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=20, font="arial 17")
reg_entry.place(x=390, y=265)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=17, font="arial 17")
date_entry.place(x=900, y=265)

Date.set(d1)

# Add the other entry widgets as per your existing UI code

# Image section
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1500, y=350)

# Set a default image (Ensure this file exists in the directory)
try:
    img = Image.open('Images/s3.jpg')  # Default image
    img = img.resize((190, 190))  # Resize to fit
    photo2 = ImageTk.PhotoImage(img)
    lbl = Label(f, bg="black", image=photo2)
    lbl.place(x=0, y=0)
    lbl.image = photo2  # Keep reference
except Exception as e:
    print("Error loading default image:", e)

#student details
obj=LabelFrame(root,text= "Student's Details",font=70,bd=2,width=1000,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=200,y=350)
Label(root,text="Full name:",font="arial 17",bg=framebg,fg=framefg).place(x=250,y=400)
Label(root,text="Date of Birth:",font="arial 17",bg=framebg,fg=framefg).place(x=250,y=450)
Label(root,text="Gender:",font="arial 17",bg=framebg,fg=framefg).place(x=250,y=500)

Label(root,text="Class:",font="arial 17",bg=framebg,fg=framefg).place(x=750,y=400)
Label(root,text="Religion:",font="arial 17",bg=framebg,fg=framefg).place(x=750,y=450)
Label(root,text="Skills:",font="arial 17",bg=framebg,fg=framefg).place(x=750,y=500)

Name=StringVar()
name_entry=Entry(obj,textvariable=Name,width=22,font="arial 17")
name_entry.place(x=200,y=30)

DOB=StringVar()
dob_entry=Entry(obj,textvariable=DOB,width=22,font="arial 17")
dob_entry.place(x=195,y=70)

radio= IntVar()
R1= Radiobutton(obj,text="Male", variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=125)

R1= Radiobutton(obj,text="Female", variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R1.place(x=230,y=125)

Religion=StringVar()
religion_entry=Entry(obj,textvariable=Religion,width=22,font="arial 17")
religion_entry.place(x=670,y=75)

Skills=StringVar()
skills_entry=Entry(obj,textvariable=Skills,width=22,font="arial 17")
skills_entry.place(x=670,y=125)

Class=Combobox(obj,values=['1','2','3','4','5','6','7','8','9','10','11','12'],font="Roboto 12",width=20,state="r")
Class.place(x=670,y=25)
Class.set("Select Class")




obj2=LabelFrame(root,text= "Parent's Details",font=70,bd=2,width=1050,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj2.place(x=200,y=670)

Label(obj2,text="Father's name:",font="arial 17",bg=framebg,fg=framefg).place(x=10,y=20)
Label(obj2,text="Occupation",font="arial 17",bg=framebg,fg=framefg).place(x=10,y=90)

F_Name=StringVar()
f_entry=Entry(obj2,textvariable=F_Name,width=20,font="arial 17")
f_entry.place(x=200,y=20)

F_Occupation=StringVar()
FO_entry=Entry(obj2,textvariable=F_Occupation,width=20,font="arial 17")
FO_entry.place(x=190,y=90)

Label(obj2,text="Mother's name:",font="arial 17",bg=framebg,fg=framefg).place(x=550,y=20)
Label(obj2,text="Occupation",font="arial 17",bg=framebg,fg=framefg).place(x=550,y=90)

M_Name=StringVar()
m_entry=Entry(obj2,textvariable=M_Name,width=20,font="arial 17")
m_entry.place(x=750,y=20)

M_Occupation=StringVar()
MO_entry=Entry(obj2,textvariable=M_Occupation,width=20,font="arial 17")
MO_entry.place(x=750,y=90)


# Buttons
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightpink", command=showimage).place(x=1500, y=590)
saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightyellow", command=Save)
saveButton.place(x=1500, y=680)
Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="purple", command=Clear).place(x=1500, y=750)
Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="orange", command=Exit).place(x=1500, y=850)

root.mainloop()
